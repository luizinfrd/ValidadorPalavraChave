import openpyxl

def verificador(planilha,  linha):
    celula_texto = planilha.cell(row=linha, column=1)
    valor_texto = celula_texto.value

    celula_situacao = planilha.cell(row=linha, column=2)


    for palavras in palavras_chaves:
        if valor_texto == None:
            print('FIM DO PROGRAMA.')
            return
        if palavras in valor_texto:
            print(f"Texto na linha {linha} é válido.")
            celula_situacao.value = 'TEXTO VÁLIDO'
            return
        
    print(f"Texto na linha {linha} é inválido.")
    celula_situacao.value = 'TEXTO INVÁLIDO'



textos = openpyxl.load_workbook('textos.xlsx')
planilha = textos.active

palavras_chaves = ['abacaxi', 'planeta', 'maça', 'natureza']

for linha in range(2, 200):
    verificador(planilha, linha)

textos.save('textos.xlsx')



